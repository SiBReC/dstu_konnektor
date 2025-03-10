import os
import sys
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from datetime import datetime
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import subprocess
import platform
import webbrowser
import shutil
import random


def resource_path(relative_path):
    """Возвращает корректный путь для ресурсов"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


class ExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Сборка журнала ДГТУ ИиВТ")
        self.files = []
        self.download_dir = os.path.join(os.getcwd(), "dstu_connector_files")
        self.remaining = 10
        os.makedirs(self.download_dir, exist_ok=True)

        self.setup_ui()
        self.create_widgets()
        self.set_icon()

    def set_icon(self):
        try:
            self.icon = resource_path("icon.ico")
            self.root.iconbitmap(self.icon)
        except Exception as e:
            print(f"Ошибка загрузки иконки: {e}")

    def setup_ui(self):
        style = ttk.Style()
        style.theme_use('clam')

        # Настройка цветовой схемы
        self.root.configure(bg='#f0f0f0')
        style.configure('TFrame', background='#f0f0f0')
        style.configure('TLabel', background='#f0f0f0', font=('Arial', 10))
        style.configure('TButton', font=('Arial', 10), padding=8,
                        background='#4a7a8c', foreground='white')
        style.map('TButton',
                  background=[('active', '#5a8a9c'), ('pressed', '#3a6a7c')])

        style.configure('Title.TLabel', font=('Arial', 14, 'bold'),
                        foreground='#2d5b6c')
        style.configure('Listbox', font=('Arial', 10), background='white')
        style.configure('Horizontal.TProgressbar', troughcolor='#e0e0e0',
                        background='#4a7a8c', thickness=20)
        style.configure('TLabelframe', background='#f0f0f0', relief='groove')
        style.configure('TLabelframe.Label', background='#f0f0f0',
                        foreground='#2d5b6c')

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding=(15, 10))
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Заголовок
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(title_frame, text="Сборка журнала ДГТУ ИиВТ",
                  style='Title.TLabel').pack(side=tk.LEFT)

        # Список файлов
        list_frame = ttk.LabelFrame(main_frame, text=" Выбранные файлы ", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.file_list = tk.Listbox(list_frame, width=60, height=10,
                                    selectmode=tk.EXTENDED, font=('Arial', 10),
                                    relief='flat', highlightthickness=1,
                                    bg='white', activestyle='none')
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL,
                                  command=self.file_list.yview)
        self.file_list.configure(yscrollcommand=scrollbar.set)

        self.file_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Кнопки управления
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        buttons = [
            ("📁 Добавить файлы", self.add_files),
            ("🧹 Очистить список", self.clear_list),
            ("🔄 Объединить", self.merge_files),
            ("🌐 Онлайн-выгрузка(BETA)", self.online_export),
            ("❓ Как пользоваться", self.show_instructions)
        ]

        for text, command in buttons:
            btn = ttk.Button(btn_frame, text=text, command=command,
                             style='TButton')
            btn.pack(side=tk.LEFT, padx=3, ipadx=5, ipady=3)

        # Статусная панель
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X, pady=(5, 0))

        self.progress = ttk.Progressbar(status_frame, orient=tk.HORIZONTAL,
                                        mode='determinate',
                                        style='Horizontal.TProgressbar')
        self.status_label = ttk.Label(status_frame, text="Готов к работе",
                                      font=('Arial', 9), foreground='#555555')
        self.status_label.pack(side=tk.LEFT)
        self.progress.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))

    def add_files(self):
        new_files = filedialog.askopenfilenames(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if new_files:
            self.files.extend(sorted(new_files))
            self.update_file_list()
            self.status_label.config(text=f"Загружено файлов: {len(self.files)}")

    def clear_list(self):
        self.files = []
        self.file_list.delete(0, tk.END)
        self.status_label.config(text="Список файлов очищен")

    def show_disclaimer(self):
        disclaimer_window = tk.Toplevel(self.root)
        disclaimer_window.title("Важное предупреждение")
        disclaimer_window.configure(bg='#f0f0f0')
        self.set_child_icon(disclaimer_window)

        text_frame = ttk.Frame(disclaimer_window, padding=10)
        text_frame.pack(fill=tk.BOTH, expand=True)

        text = tk.Text(text_frame, wrap=tk.WORD, width=60, height=10,
                       font=('Arial', 10), bg='white', relief='flat')
        scrollbar = ttk.Scrollbar(text_frame, command=text.yview)
        text.configure(yscrollcommand=scrollbar.set)

        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        disclaimer_text = (
            "ВНИМАНИЕ!\n\n"
            "Разработчик не несёт ответственности за сохранность ваших введенных данных, "
            "но уведомляет, что оригинальная версия программы НЕ передает данные никому, "
            "кроме сайта https://edu.donstu.ru. Данные передаются единоразово и доступны "
            "только Вам и этому сайту. Оригинальность данного дистрибутива можно посмотреть "
            "в окне 'Как пользоваться'.\n\n"
            "Поддержка BETA функций НЕ гарантирована!\n"
            "Любая ошибка BETA(функций) фиксится перезагрузкой программы"
        )
        text.insert(tk.END, disclaimer_text)
        text.config(state=tk.DISABLED)

        btn_frame = ttk.Frame(disclaimer_window)
        btn_frame.pack(pady=10)

        ttk.Button(btn_frame, text="Я понимаю", style='TButton',
                   command=lambda: [disclaimer_window.destroy(), self.show_auth_window()]
                   ).pack()

    def show_auth_window(self):
        auth_window = tk.Toplevel(self.root)
        auth_window.title("Авторизация в системе ДГТУ")
        auth_window.resizable(False, False)
        auth_window.configure(bg='#f0f0f0')
        self.set_child_icon(auth_window)

        form_frame = ttk.Frame(auth_window, padding=15)
        form_frame.pack()

        ttk.Label(form_frame, text="Логин:").grid(row=0, column=0, padx=5, pady=5)
        self.login_entry = ttk.Entry(form_frame, width=25, font=('Arial', 10))
        self.login_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(form_frame, text="Пароль:").grid(row=1, column=0, padx=5, pady=5)
        self.password_entry = ttk.Entry(form_frame, show="*", width=25, font=('Arial', 10))
        self.password_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Button(form_frame, text="Начать выгрузку", style='TButton',
                   command=self.start_online_export).grid(row=2, columnspan=2, pady=15)

    def check_chrome_installed(self):
        """Проверяет наличие Google Chrome в системе с учетом различных путей"""
        # Проверка через shutil.which
        chrome_names = ['google-chrome', 'chrome', 'google-chrome-stable', 'chromium']
        for name in chrome_names:
            if shutil.which(name):
                return True

        # Дополнительные проверки для конкретных ОС
        system = platform.system()
        if system == "Windows":
            common_paths = [
                os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
                os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
                os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe")
            ]
            for path in common_paths:
                if os.path.exists(path):
                    return True

        elif system == "Darwin":
            mac_paths = [
                "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
                "/Applications/Chromium.app/Contents/MacOS/Chromium"
            ]
            for path in mac_paths:
                if os.path.exists(path):
                    return True

        elif system == "Linux":
            linux_paths = [
                "/usr/bin/google-chrome",
                "/usr/local/bin/google-chrome",
                "/opt/google/chrome/chrome",
                "/snap/bin/chromium"
            ]
            for path in linux_paths:
                if os.path.exists(path):
                    return True

        return False

    def show_chrome_install_dialog(self):
        """Показывает диалог установки Chrome"""
        msg = (
            "Для работы онлайн-выгрузки требуется Google Chrome.\n"
            "Установить его сейчас?"
        )
        answer = messagebox.askyesno(
            "Требуется установка",
            msg,
            parent=self.root
        )

        if answer:
            try:
                # Открываем страницу загрузки Chrome
                webbrowser.open("https://www.google.ru/chrome/")
                # Показываем кнопку для повторной проверки
                retry_answer = messagebox.askyesno(
                    "Проверка установки",
                    "После установки Chrome нажмите 'Да' для повторной проверки",
                    parent=self.root
                )
                if retry_answer:
                    return self.check_chrome_installed()
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось открыть страницу загрузки: {e}")
        return False

    def online_export(self):
        """Обновленный метод для онлайн-выгрузки"""
        if not self.check_chrome_installed():
            if not self.show_chrome_install_dialog():
                messagebox.showinfo(
                    "Отмена",
                    "Онлайн-выгрузка невозможна без Google Chrome",
                    parent=self.root
                )
                return
        self.show_disclaimer()

    def start_online_export(self):

        login = self.login_entry.get()
        password = self.password_entry.get()

        if not login or not password:
            messagebox.showwarning("Ошибка", "Введите логин и пароль!")
            return

        driver = None
        try:
            self.clean_downloads_folder()

            options = webdriver.ChromeOptions()
            prefs = {
                "download.default_directory": self.download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True
            }
            options.add_experimental_option("prefs", prefs)
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")

            self.status_label.config(text="Инициализация браузера...")
            self.progress.pack(pady=5)

            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            driver.set_page_load_timeout(30)

            # Авторизация
            self.status_label.config(text="Авторизация на сайте...")
            driver.get("https://edu.donstu.ru/WebApp/#/login")

            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.NAME, "login"))
            ).send_keys(login)

            driver.find_element(By.NAME, "password").send_keys(password)
            driver.find_element(By.CSS_SELECTOR, "button.stud").click()

            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".v-main__wrap")))
                self.status_label.config(text="Авторизация успешна!")
                self.progress['value'] = 25
            except TimeoutException:
                messagebox.showerror("Ошибка", "Неверные учетные данные")
                return

            # Сбор журналов
            # Сбор журналов
            self.status_label.config(text="Поиск журналов...")
            try:
                # Улучшенная обработка навигации
                def is_page_loaded(driver):
                    return driver.execute_script("return document.readyState") == "complete"

                # Очистка кук и кэша перед загрузкой
                driver.delete_all_cookies()

                # Улучшенный запрос с повторными попытками
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        driver.get("https://edu.donstu.ru/WebApp/#/Journals/JournalList")
                        # Добавленный код для клика по вкладке "Все"
                        WebDriverWait(driver, 15).until(
                            EC.presence_of_element_located((By.XPATH, "//div[@role='tab' and contains(., 'Все')]"))
                        ).click()
                        time.sleep(2)  # Ожидание 3 секунды после клика
                        # Комбинированное ожидание
                        WebDriverWait(driver, 15).until(
                            lambda d: is_page_loaded(d) and
                                      d.find_elements(By.CSS_SELECTOR, "a[href^='#/Journals/Journal/']")
                        )
                        break
                    except TimeoutException:
                        if attempt == max_retries - 1:
                            raise
                        driver.refresh()
                        time.sleep(2)

                # Дополнительная проверка загрузки
                if "JournalList" not in driver.current_url:
                    raise Exception("Не удалось подтвердить успешную навигацию")

                # Добавляем случайные задержки для имитации человеческого поведения
                time.sleep(random.uniform(0.5, 1.5))

                # Проверка на наличие CAPTCHA или блокировки
                if "Доступ запрещен" in driver.page_source:
                    raise Exception("Обнаружена блокировка доступа")

            except Exception as e:
                error_msg = f"Ошибка загрузки страницы: {str(e)}"

                # Проверка интернет-соединения
                try:
                    subprocess.check_call(["ping", "-c", "1", "8.8.8.8"])
                except subprocess.CalledProcessError:
                    error_msg += "\nПроверьте интернет-соединение"

                messagebox.showerror("Критическая ошибка", error_msg)
                return


            journals = driver.find_elements(By.CSS_SELECTOR, "a[href^='#/Journals/Journal/']")
            journal_urls = list({j.get_attribute("href") for j in journals})

            if not journal_urls:
                messagebox.showinfo("Информация", "Нет доступных журналов для выгрузки")
                return

            total = len(journal_urls)
            self.progress['maximum'] = total

            for idx, url in enumerate(journal_urls, 1):
                try:
                    self.status_label.config(text=f"Обработка журнала {idx}/{total}")
                    driver.get(url)
                    time.sleep(2)

                    export_btn = WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, ".dx-icon-export-excel-button"))
                    )
                    export_btn.click()
                    time.sleep(3)
                    self.progress['value'] = idx

                except Exception as e:
                    messagebox.showwarning("Ошибка", f"Ошибка при экспорте: {str(e)}")
                    continue

            self.add_files_from_downloads()
            self.remaining = 10
            self.update_timer()

        except Exception as e:
            messagebox.showerror("Критическая ошибка", f"Ошибка: {str(e)}")
        finally:
            if driver is not None:
                driver.quit()
            self.progress.pack_forget()

    def update_timer(self):
        if self.remaining > 0:
            self.status_label.config(
                text=f"Ожидайте начало объединения... {self.remaining} сек."
            )
            self.remaining -= 1
            self.root.after(1000, self.update_timer)
        else:
            self.merge_files()

    def clean_downloads_folder(self):
        for filename in os.listdir(self.download_dir):
            file_path = os.path.join(self.download_dir, filename)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось удалить {file_path}: {e}")

    def add_files_from_downloads(self):
        try:
            new_files = [os.path.join(self.download_dir, f)
                         for f in os.listdir(self.download_dir)
                         if f.endswith('.xlsx')]

            if not new_files:
                messagebox.showwarning("Внимание", "Не найдено Excel-файлов")
                return

            self.files = sorted(new_files)
            self.update_file_list()
            self.status_label.config(text=f"Добавлено {len(new_files)} файлов")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка добавления файлов: {e}")

    def merge_files(self):
        if not self.files:
            messagebox.showwarning("Ошибка", "Добавьте файлы для объединения")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Выберите место для сохранения объединенного файла"
        )

        if not save_path:
            return

        try:
            new_wb = Workbook()
            new_wb.properties.creator = "DSTU_Коннектор"
            if new_wb.sheetnames:
                new_wb.remove(new_wb.active)

            for file_path in sorted(self.files):
                try:
                    wb = load_workbook(file_path, data_only=False)
                    ws = wb.worksheets[0]
                    original_name = os.path.basename(file_path)
                    sheet_name = self.process_sheet_name(original_name)

                    unique_name = sheet_name
                    counter = 1
                    while unique_name in new_wb.sheetnames:
                        unique_name = f"{sheet_name}_{counter}"
                        if len(unique_name) > 31:
                            max_base_length = 31 - len(str(counter)) - 1
                            unique_name = f"{sheet_name[:max_base_length]}_{counter}"
                        counter += 1

                    new_ws = new_wb.create_sheet(title=unique_name)

                    for row in ws.iter_rows():
                        for cell in row:
                            new_cell = new_ws.cell(
                                row=cell.row,
                                column=cell.column,
                                value=cell.value
                            )
                            if cell.has_style:
                                new_cell.font = cell.font.copy()
                                new_cell.border = cell.border.copy()
                                new_cell.fill = cell.fill.copy()
                                new_cell.number_format = cell.number_format
                                new_cell.protection = cell.protection.copy()
                                new_cell.alignment = cell.alignment.copy()

                    for merged_range in ws.merged_cells.ranges:
                        new_ws.merge_cells(str(merged_range))

                    for col in ws.column_dimensions:
                        new_ws.column_dimensions[col].width = ws.column_dimensions[col].width

                    for row in ws.row_dimensions:
                        new_ws.row_dimensions[row].height = ws.row_dimensions[row].height

                    wb.close()
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Ошибка обработки файла {file_path}:\n{str(e)}")
                    continue

            new_wb.save(save_path)
            new_wb.close()
            self.status_label.config(text="Готово")
            messagebox.showinfo("Готово", f"Файлы успешно объединены!\nСохранено в:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при объединении: {e}")

    def process_sheet_name(self, filename):
        name = os.path.splitext(filename)[0]
        name = re.sub(r'(?i)журнал|[()]', '', name)
        name = re.sub(r'\s*,\s*', ',', name)
        name = re.sub(r',+', ',', name)

        for pattern in ["Лаб,", "Лек,", "Пр,"]:
            if pattern in name:
                name = name.split(pattern, 1)[0] + pattern.rstrip(',')
                break

        parts = re.split(r'[\s_,]+', name)
        processed = []
        for part in parts:
            if not part:
                continue
            if part.lower() == 'п_г':
                processed.append('п_г')
                continue
            trunc_len = 5 if len(part) > 10 else 3
            processed.append(part[:trunc_len])

        clean_name = '_'.join(processed)
        clean_name = re.sub(r'[\\/*?[\]:,.]', '_', clean_name)
        clean_name = re.sub(r'_+', '_', clean_name).strip('_')

        return clean_name[:31]

    def update_file_list(self):
        self.file_list.delete(0, tk.END)
        for file in self.files:
            self.file_list.insert(tk.END, os.path.basename(file))

    def show_instructions(self):
        instructions_window = tk.Toplevel(self.root)
        instructions_window.title("Инструкция")
        instructions_window.configure(bg='#f0f0f0')
        self.set_child_icon(instructions_window)

        text_frame = ttk.Frame(instructions_window, padding=10)
        text_frame.pack(fill=tk.BOTH, expand=True)

        text = tk.Text(text_frame, wrap=tk.WORD, width=60, height=15,
                       font=('Arial', 10), bg='white', relief='flat')
        scrollbar = ttk.Scrollbar(text_frame, command=text.yview)
        text.configure(yscrollcommand=scrollbar.set)

        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        instructions_text = (
            "1) Для ручного использования:\n"
            "   - Нажмите 'Добавить файлы' и выберите XLSX файлы\n"
            "   - Нажмите 'Объединить' и выберите место сохранения\n\n"
            "2) Для онлайн-выгрузки:\n"
            "   - Нажмите 'Онлайн-выгрузка'\n"
            "   - Введите логин/пароль от edu.donstu.ru\n"
            "   - После загрузки ожидайте 10 секунд\n"
            "   - Выберите место сохранения объединенного файла\n\n"
            "3) Очистка списка - кнопка 'Очистить список'\n\n"
            "Оригинальный HASH программы расположен по адресу:\n"
            "https://disk.yandex.ru/i/FMvuS8UOpkmUSw\n"
            "Время создания оригинального файла: 10.03.25 15:02 МСК\n\n"
            "Разработчик tg:@FIRTUX\nВерсия: v0.0.5-100325\n"
            "Буду рад любому виду поддержки!\n\n"
            "Данная программа поддерживается до 01.09.2025\n"
            "Поддержка BETA функций НЕ гарантирована!"
        )
        text.insert(tk.END, instructions_text)
        text.config(state=tk.DISABLED)

        btn_frame = ttk.Frame(instructions_window)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="OK", style='TButton',
                   command=instructions_window.destroy).pack()

    def set_child_icon(self, window):
        try:
            window.iconbitmap(self.icon)
        except Exception as e:
            print(f"Ошибка загрузки иконки: {e}")


def check_activation():
    current_date = datetime.now().date()
    cutoff_date = datetime(2025, 9, 1).date()
    if current_date >= cutoff_date:
        root = tk.Tk()
        root.withdraw()
        try:
            root.iconbitmap(resource_path("icon.ico"))
            root.iconbitmap(icon_path)
        except:
            pass
        key = simpledialog.askstring("Активация", "Поддержка данной версии программы завершена! Обратитесь к разработчику или введите ключ отладки:", show='*')
        if key != "*":
            root.destroy()
            return False
        root.destroy()
    return True


if __name__ == "__main__":
    if not check_activation():
        exit()
    root = tk.Tk()
    try:
        icon_path = resource_path("icon.ico")
        root.iconbitmap(icon_path)
    except Exception as e:
        print(f"Ошибка загрузки иконки: {e}")
    app = ExcelMergerApp(root)
    root.mainloop()
